import sys
import os

from pathlib import Path

from math import ceil, isclose

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

import pandas as pd
import matplotlib.pyplot as plt

from PySide6.QtWidgets import QApplication, QMainWindow, QDialog, QMessageBox, QFileDialog
from PySide6.QtGui import QPixmap, QPainter
from PySide6.QtPrintSupport import QPrinter, QPrintDialog
from PySide6.QtCore import Qt

from MainWindow_ui import Ui_MainWindow  # Интерфейс главного окна
from InfoWindow_ui import Ui_info_win  # Интерфейс окна с информацией


# Инициализация окна "Инфо"
class InfoWindow(QDialog):
    def __init__(self):
        super().__init__()
        self.ui = Ui_info_win()
        self.ui.setupUi(self)  # Инициализация интерфейса окна "Инфо"

        # Привязка кнопки закрытия окна к методу close
        self.ui.pushButton.clicked.connect(self.close)


# Инициализация главного окна
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()  # Инициализация интерфейса главного окна
        self.ui.setupUi(self)

        # Привязка кнопок к соответствующим методам
        self.ui.button_info.clicked.connect(self.open_info_window)  # Кнопка открытия окна "Инфо"
        self.ui.button_start.clicked.connect(self.generate_graph)  # Кнопка генерации графика
        self.ui.button_reset.clicked.connect(self.restart_graph)  # Кнопка сброса графика
        self.ui.button_print.clicked.connect(self.print_graph)  # Кнопка печати графика
        self.ui.button_excel.clicked.connect(self.open_excel)  # Кнопка открытия Excel-файла
        self.ui.pushButton.clicked.connect(self.show_previous_graph)  # Кнопка предыдущего графика
        self.ui.pushButton_2.clicked.connect(self.show_next_graph)  # Кнопка следующего графика
        self.ui.button_upload.clicked.connect(self.upload_excel)  # Кнопка загрузки Excel-файла

        # Отключение некоторых кнопок до генерации графика
        self.ui.button_reset.setEnabled(False)
        self.ui.button_reset.setEnabled(False)
        self.ui.button_print.setEnabled(False)
        self.ui.button_excel.setEnabled(False)

        # Скрытие кнопок для переключения графиков до их генерации
        self.ui.pushButton.setHidden(True)
        self.ui.pushButton_2.setHidden(True)

        # Привязка проверки заполнения полей ввода к методу check_input_fields
        self.ui.entry_t0.textChanged.connect(self.check_input_fields)
        self.ui.entry_tk.textChanged.connect(self.check_input_fields)
        self.ui.entry_dt1.textChanged.connect(self.check_input_fields)
        self.ui.entry_dt2.textChanged.connect(self.check_input_fields)
        self.ui.entry_x0.textChanged.connect(self.check_input_fields)
        self.ui.entry_v0.textChanged.connect(self.check_input_fields)
        self.ui.entry_k.textChanged.connect(self.check_input_fields)
        self.ui.entry_m.textChanged.connect(self.check_input_fields)
        self.ui.entry_y.textChanged.connect(self.check_input_fields)

        # Переменные для хранения информации о графиках
        self.info_window = None  # Окно с информацией
        self.pixmap = None  # Изображение для отображения

        # Счётчик графиков и пути к ним
        self.graph_counter = 0  # Счётчик графиков
        self.graph_paths = []  # Пути к сгенерированным графикам
        self.excel_paths = []  # Пути к сгенерированным Excel-файлам

    # Обновление графика при изменении размера окна
    def resizeEvent(self, event):
        try:
            self.update_graph()  # Обновляем график
        except:
            pass
        super().resizeEvent(event)

    # Открытие окна "Инфо"
    def open_info_window(self):
        if self.info_window is None:
            self.info_window = InfoWindow()  # Создание окна, если его еще нет
        self.info_window.show()  # Отображение окна

    # Проверка, заполнены ли все поля ввода
    def check_input_fields(self):
        # Если хотя бы одно поле заполнено, включаем кнопку сброса
        if (self.ui.entry_t0.text() or self.ui.entry_tk.text() or self.ui.entry_dt1.text() or self.ui.entry_dt2.text()
                or self.ui.entry_x0.text() or self.ui.entry_v0.text() or self.ui.entry_k.text()
                or self.ui.entry_m.text() or self.ui.entry_y.text()):
            self.ui.button_reset.setEnabled(True)
        else:
            self.ui.button_reset.setEnabled(False)

    # Отображение ошибки при некорректных данных
    def error_incorrect_input_message(self, title, message):
        error_dialog = QMessageBox()  # Диалоговое окно для ошибки
        icon = QPixmap(Path("images/error.svg"))  # Установка иконки ошибки
        error_dialog.setWindowIcon(icon)
        error_dialog.setIcon(QMessageBox.Critical)  # Установка иконки критической ошибки
        error_dialog.setWindowTitle(title)  # Заголовок окна
        error_dialog.setText(message)  # Сообщение ошибки
        error_dialog.setStandardButtons(QMessageBox.Ok)  #Создание в окне кнопки "ОК" для выхода
        error_dialog.exec()  # Отображение диалогового окна

    # Метод генерации графика на основе данных
    def generate_graph(self):
        try:
            # Чтение данных из полей ввода
            t0 = float(self.ui.entry_t0.text().replace(',', '.'))  # начальный момент времени
            tk = float(self.ui.entry_tk.text().replace(',', '.'))  # конечный момент времени
            x0 = float(self.ui.entry_x0.text().replace(',', '.'))  # начальное положение груза
            v0 = float(self.ui.entry_v0.text().replace(',', '.'))  # начальная скорость
            m = float(self.ui.entry_m.text().replace(',', '.'))  # масса груза
            k = float(self.ui.entry_k.text().replace(',', '.'))  # коэффициент жесткости
            y = float(self.ui.entry_y.text().replace(',', '.'))  # коэффициент трения
            dt1 = float(self.ui.entry_dt1.text().replace(',', '.'))  # период расчетов (вывод значений)
            dt2 = float(self.ui.entry_dt2.text().replace(',', '.'))  # параметр дискретизации

            # Проверка введенных данных на корректность
            if t0 >= tk or dt1 <= 0 or dt2 <= 0 or dt1 < dt2 or y < 0 or m < 0 or k < 0:
                self.error_incorrect_input_message("Ошибка ввода данных",
                                                   "Убедитесь, что введённые данные соответствуют"
                                                   " следующим условиям: "
                                                   "t0 < tk; "
                                                   "dt1 > 0; dt2 > 0; dt1 >= dt2; "
                                                   "γ >= 0; m > 0; k > 0.")
                return

        except ValueError:
            # Ошибка, если данные введены некорректно
            self.error_incorrect_input_message("Ошибка ввода данных",
                                               "Убедитесь, что все поля заполнены корректно "
                                               "(введены только цифры"
                                               " и цифры через запятую или точку).")
            return

        # Рассчитываем количество шагов для временных интервалов
        # ceil() используется для округления результата вверх, чтобы гарантировать, что мы захватим весь интервал
        N = ceil((tk - t0) / dt1)
        M = ceil(dt1 / dt2)

        # Инициализация массивов для времени, скорости, положения и энергии
        t = list(False for _ in range(0, N * M + 2, 1))  # Массив времени
        x = list(False for _ in range(0, N * M + 2, 1))  # Массив положения
        v = list(False for _ in range(0, N * M + 2, 1))  # Массив скорости
        e = list(False for _ in range(0, N * M + 2, 1))   # Массив энергии (если трение в системе отсутствует)

        try:
            # Установка начальных условий
            x[0] = x0
            v[0] = v0
            t[0] = t0
            # Начальное значение полной энергии
            e[0] = (m * v[0] ** 2 + k * x[0] ** 2) / 2

            index = 1

            # Численное решение уравнения методом Эйлера
            for i in range(1, N + 1):  # Перебор шагов по сетке времени dt1
                for j in range(1, M + 1):  # Вложенный цикл для временной сетки dt2
                    # Корректировка времени для шагов по dt2
                    if t[index - 1] + dt2 > t0 + dt1 * i:
                        t[index] = t0 + dt1 * i
                    else:
                        t[index] = t[index - 1] + dt2

                    # Остановка, если текущее время превышает конечное время
                    if t[index] > tk:
                        t[index] = tk

                    # Расчёт скорости и положения методом Эйлера
                    v[index] = v[index - 1] + (-(k / m if m > 0 else 0) * x[index - 1] - y * v[index - 1]) * dt2
                    x[index] = x[index - 1] + v[index] * dt2

                    # Расчёт полной энергии на каждом шаге, если время совпадает с целым шагом по dt1
                    if isclose(t[index], t0 + dt1 * i, rel_tol=0.001):
                        e[index] = (m * v[index] ** 2 + k * x[index] ** 2) / 2

                    # Прерывание цикла, если достигнуто конечное время
                    if t[index] == tk:
                        break

                    # Увеличение индекса для перехода к следующему шагу
                    index += 1

            # Если конечное время tk не оказалось в массиве t, то оно добавляется вручную
            if tk not in t:
                while t[-1] is False:  # Удаление лишних элементов с конца массива
                    del t[-1]
                    del v[-1]
                    del x[-1]
                    del e[-1]
                t.append(tk)  # Добавление tk в массив времени
                # Расчёт последней скорости и положения на основе конечного времени
                v.append(v[-1] + (-(k / m if m > 0 else 0) ** 2 * x[-1] - y * v[-1]) * dt2)  # Расчёт последней скорости
                x.append(x[-1] + v[-1] * dt2)  # Расчёт последнего положения
                e.append((m * v[-1] ** 2 + k * x[-1] ** 2) / 2)  # Расчёт последней энергии
            else:
                # Если tk присутствует в массиве t несколько раз, оставляем только одно вхождение
                count = t.count(tk)
                if count == 1:
                    while t[-1] != tk:
                        del t[-1]
                        del v[-1]
                        del x[-1]
                        del e[-1]
                else:
                    flag = True
                    while flag:
                        if t[-1] == tk:
                            count -= 1

                        del t[-1]
                        del v[-1]
                        del x[-1]
                        del e[-1]

                        if count == 1:
                            flag = False
        except Exception:
            # Обработка ошибок и вывод сообщения об ошибке при некорректных входных данных
            self.error_incorrect_input_message("Ошибка обработки", "Входные параметры приводят"
                                                                   " к слишком большим числам. "
                                                                   "Попробуйте изменить вводимые числа.")
            return

        # Проверка, существует ли директория для сохранения графиков, если нет - создается
        if not os.path.exists("graph"):
            os.makedirs("graph")

        # Функция для оформления заголовков в Excel (применение жирного шрифта и тонкой границы)
        def header_preset(ws, columns):
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            # Применение жирного шрифта и тонкой границы к заголовкам
            for col in columns:
                ws[col].font = Font(bold=True)
                ws[col].border = thin_border

            # Автоматическая подстройка ширины столбцов
            for column_cells in ws.columns:
                max_length = max(len(str(cell.value)) for cell in column_cells)
                adjusted_width = max_length + 2
                column_letter = get_column_letter(column_cells[0].column)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Функция для создания Excel файла с заданным стилем для заголовков
        def create_excel_with_preset(file_path, data, columns, M):
            df = pd.DataFrame(data)  # Создание DataFrame из данных
            df.to_excel(file_path, index=False)  # Сохранение данных в Excel файл
            wb = load_workbook(file_path)  # Открытие файла
            ws = wb.active  # Выбор активного листа
            last_row = ws.max_row  # Определение последней строки
            header_preset(ws, columns)  # Применение стилей к заголовкам

            # Заливка строк разными цветами для выделения данных
            filler1 = PatternFill(start_color='FFCBDB', end_color='FFCBDB', fill_type='solid')
            filler2 = PatternFill(start_color='FFC8A8', end_color='FFC8A8', fill_type='solid')
            filler3 = PatternFill(start_color='BDECB6', end_color='BDECB6', fill_type='solid')
            for row in range(2, last_row + 1):
                if (row - 2) % M == 0:
                    if file_path == excel_file_path_1 and y == 0:
                        # Оформление для определенных файлов и значений трения
                        ws[f'A{row}'].fill = filler1
                        ws[f'B{row}'].fill = filler1
                        ws[f'C{row}'].fill = filler1
                        ws[f'D{row}'].fill = filler2
                        ws['E2'].fill = filler3
                        ws[f'F{row}'].fill = filler2
                        ws['G2'].fill = filler3
                    elif file_path == excel_file_path_1:
                        ws[f'A{row}'].fill = filler1
                        ws[f'B{row}'].fill = filler1
                        ws[f'C{row}'].fill = filler1
                    else:
                        ws[f'A{row}'].fill = filler1
                        ws[f'B{row}'].fill = filler1
            # Сохранение изменений в файл
            wb.save(file_path)
            wb.close()

        # Создание графиков и сохранение их в файлы
        # График x(t) и v(t)
        fig1, ax1 = plt.subplots(figsize=(10, 6))
        fig1.patch.set_facecolor((1, 1, 1, 0))  # Установка прозрачного фона для графика

        # Построение графика x(t) - синяя линия
        ax1.plot(t, x, label="x(t) - Положение", color="blue")

        # Построение графика v(t) - розовая линия
        ax1.plot(t, v, label="v(t) - Скорость", color="pink")

        # Установка заголовка и меток осей
        ax1.set_title("Колебания линейного гармонического осциллятора с учётом трения")
        ax1.set_xlabel("t - Время")
        ax1.set_ylabel("x - Положение; v - Скорость")

        # Включение сетки и легенды
        ax1.grid(True)
        ax1.legend()

        # Сохранение графика x(t) и v(t) в файл с прозрачным фоном
        graph_path_1 = Path("graph/oscillator_graph_xvt.png")
        plt.savefig(graph_path_1, transparent=True)
        plt.close()  # Закрываем фигуру после сохранения

        # Создание Excel-файла для графика xvt
        excel_file_path_1 = Path("graph/oscillator_data_xvt.xlsx")
        if y == 0:
            # Если трение y равно 0, добавляем столбец с полной энергией и её расчёт
            data = {"Время t": t, "Положение x": x, "Скорость v": v,
                    "Полная энергия Ei": [info if info else -2 for info in e]}
            create_excel_with_preset(excel_file_path_1, data, ["A1", "B1", "C1", "D1"], M)
        else:
            # Если трение y не равно 0, создаем Excel-файл без столбца полной энергии
            data = {"Время t": t, "Положение x": x, "Скорость v": v}
            create_excel_with_preset(excel_file_path_1, data, ["A1", "B1", "C1"], M)

        # Дополнительные операции, если трение y == 0
        if y == 0:
            wb = load_workbook(excel_file_path_1)  # Открываем созданный Excel-файл
            ws = wb.active  # Выбираем активный лист
            last_row = ws.max_row  # Получаем номер последней строки

            # Очистка значений в колонке "Полная энергия Ei", если индекс не кратен M
            for cell in range(2, last_row + 1):
                if (cell - 2) % M != 0:
                    ws[f"D{cell}"].value = ""

            # Добавление заголовков и формул для расчета энергии и погрешностей
            ws['E1'] = "Среднее значение полной энергии Esr"
            ws['E2'] = f"=AVERAGE(D2:D{last_row})"  # Формула для среднего значения энергии
            ws['F1'] = "Относительная погрешность по энергии di"
            for row in range(2, last_row + 1):
                if (row - 2) % M == 0:
                    ws[f'F{row}'] = f"=ABS($E$2 - $D{row}) / ABS($E$2)"  # Формула для относительной погрешности
            ws['G1'] = "Среднеквадратичная относительная погрешность dsr"
            ws['G2'] = f"=SQRT(1 / {N} * SUMPRODUCT(($F$2:$F${last_row})^2))"

            # Применение стиля к заголовкам
            header_preset(ws, ["A1", "B1", "C1", "D1", "E1", "F1", "G1"])

            wb.save(excel_file_path_1)  # Сохраняем изменения в файл
            wb.close()  # Закрываем файл

        # Создание Excel-файлов для графика x(t)
        excel_file_path_2 = Path("graph/oscillator_data_xt.xlsx")
        data = {"Время t": t, "Положение x": x}
        create_excel_with_preset(excel_file_path_2, data, ["A1", "B1"], M)

        # Создание Excel-файлов для графика v(t)
        excel_file_path_3 = Path("graph/oscillator_data_vt.xlsx")
        data = {"Время t": t, "Скорость v": v}
        create_excel_with_preset(excel_file_path_3, data, ["A1", "B1"], M)

        # Создание графика x(t)
        fig2, ax2 = plt.subplots(figsize=(10, 6))  # Создаем фигуру и ось для второго графика
        fig2.patch.set_facecolor((1, 1, 1, 0))  # Устанавливаем прозрачный фон для фигуры

        # Построение графика x(t) - синяя линия
        ax2.plot(t, x, label="x(t) - Положение", color="blue")

        ax2.set_title("Колебания линейного гармонического осциллятора с учётом трения")
        ax2.set_xlabel("t - Время")
        ax2.set_ylabel("x - Положение")
        ax2.grid(True)
        ax2.legend()

        # Сохранение графика xt в файл с прозрачным фоном
        graph_path_2 = Path("graph/oscillator_graph_xt.png")
        plt.savefig(graph_path_2, transparent=True)
        plt.close()  # Закрываем фигуру после сохранения

        # Создание графика v(t)
        fig3, ax3 = plt.subplots(figsize=(10, 6))  # Создаем фигуру и ось для третьего графика
        fig3.patch.set_facecolor((1, 1, 1, 0))  # Устанавливаем прозрачный фон для фигуры

        # Построение графика v(t) - розовая линия
        ax3.plot(t, v, label="v(t) - Скорость", color="pink")

        ax3.set_title("Колебания линейного гармонического осциллятора с учётом трения")
        ax3.set_xlabel("t - Время")
        ax3.set_ylabel("v - Скорость")
        ax3.grid(True)
        ax3.legend()

        # Сохранение графика v(t) в файл с прозрачным фоном
        graph_path_3 = Path("graph/oscillator_graph_vt.png")
        plt.savefig(graph_path_3, transparent=True)
        plt.close()  # Закрываем фигуру после сохранения

        # Сохранение путей к графикам и Excel-файлам в списки для дальнейшего использования
        self.graph_paths = [graph_path_1, graph_path_2, graph_path_3]
        self.excel_paths = [excel_file_path_1, excel_file_path_2, excel_file_path_3]

        # Установка начального изображения графика (индекс 0)
        self.graph_counter = 0
        self.update_graph()  # Обновляем отображение графика

        # Делаем видимыми кнопки для переключения графиков
        self.ui.pushButton.setHidden(False)  # Кнопка "Назад"
        self.ui.pushButton.setEnabled(False)  # Отключаем кнопку "Назад", если это первый график
        self.ui.pushButton_2.setHidden(False)  # Кнопка "Вперед"

        # Разблокировка кнопок "Сброс", "Печать", "Excel"
        self.ui.button_reset.setEnabled(True)
        self.ui.button_print.setEnabled(True)
        self.ui.button_excel.setEnabled(True)

    def show_next_graph(self):
        # Переключение на следующий график
        if self.graph_counter < len(self.graph_paths) - 1:
            self.graph_counter += 1
            self.update_graph()  # Обновляем отображение нового графика

        # Если достигли последнего графика, отключаем кнопку "Вперед"
        if self.graph_counter == len(self.graph_paths) - 1:
            self.ui.pushButton_2.setEnabled(False)

        # Всегда включаем кнопку "Назад", если это не первый график
        if self.graph_counter > 0:
            self.ui.pushButton.setEnabled(True)

    def show_previous_graph(self):
        # Переключение на предыдущий график
        if self.graph_counter > 0:
            self.graph_counter -= 1
            self.update_graph()  # Обновляем отображение нового графика

        # Если достигли первого графика, отключаем кнопку "Назад"
        if self.graph_counter == 0:
            self.ui.pushButton.setEnabled(False)

        # Всегда включаем кнопку "Вперед", если это не последний график
        if self.graph_counter < len(self.graph_paths) - 1:
            self.ui.pushButton_2.setEnabled(True)

    def update_graph(self):
        # Обновляем изображение текущего графика и отображаем его
        self.pixmap = QPixmap(self.graph_paths[self.graph_counter])

        graph_size = self.ui.label_graph.size()  # Получаем размер виджета для отображения графика

        # Масштабируем изображение, сохраняя пропорции
        scaled_pixmap = self.pixmap.scaled(graph_size, Qt.KeepAspectRatio, Qt.SmoothTransformation)

        # Устанавливаем масштабированное изображение в QLabel
        self.ui.label_graph.setPixmap(scaled_pixmap)

    def restart_graph(self):
        # Очистка полей ввода и графика
        self.ui.entry_t0.clear()
        self.ui.entry_tk.clear()
        self.ui.entry_dt1.clear()
        self.ui.entry_dt2.clear()
        self.ui.entry_x0.clear()
        self.ui.entry_v0.clear()
        self.ui.entry_k.clear()
        self.ui.entry_m.clear()
        self.ui.entry_y.clear()

        # Очищение лейбла с графиком
        self.ui.label_graph.clear()

        # Блокировка кнопок "Печать", "Excel", разблокировка кнопки "Сгенерировать", скрытие стрелок переключения
        self.ui.button_print.setEnabled(False)
        self.ui.button_excel.setEnabled(False)
        self.ui.button_start.setEnabled(True)
        self.ui.pushButton.setHidden(True)
        self.ui.pushButton_2.setHidden(True)

        # Сброс pixmap для предотвращения случайной печати старого изображения
        self.pixmap = None

    def print_graph(self):
        # Проверяем, что pixmap был создан (график существует)
        if self.pixmap:
            # Создаем объект QPrinter для настройки параметров печати
            printer = QPrinter()

            # Открываем диалог печати с использованием QPrintDialog,
            # предоставляя ему объект QPrinter для отображения диалога
            dialog = QPrintDialog(printer, self)

            # Если пользователь подтверждает диалог печати
            if dialog.exec():
                # Создаем объект QPainter, который будет использоваться для рисования на принтере
                painter = QPainter(printer)

                # Получаем прямоугольник видимой области для рисования на принтере
                rect = painter.viewport()

                # Получаем размер pixmap, который будет напечатан
                size = self.pixmap.size()

                # Масштабируем размер pixmap так, чтобы он вписывался в область видимости принтера,
                # сохраняя пропорции изображения
                size.scale(rect.size(), Qt.KeepAspectRatio)

                # Устанавливаем видимую область для рисования на принтере
                painter.setViewport(rect.x(), rect.y(), size.width(), size.height())

                # Устанавливаем окно для рисования на pixmap
                painter.setWindow(self.pixmap.rect())

                # Рисуем pixmap в указанной области
                painter.drawPixmap(0, 0, self.pixmap)

                # Завершаем работу с QPainter
                painter.end()

    def open_excel(self):
        # Открытие Excel файла, связанного с текущим графиком
        excel_file_path = self.excel_paths[self.graph_counter]
        if os.path.exists(excel_file_path):
            os.startfile(excel_file_path)  # Открываем файл с помощью стандартного приложения
        else:
            # Отображаем сообщение об ошибке, если файл не найден
            self.error_incorrect_input_message("Ошибка при открытии файла", "Файл не найден в директории.")

    def upload_excel(self):
        # Открытие проводника для выбора Excel-файла
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "Выберите Excel-файл", "",
                                                   "Excel Files (*.xlsx);;All Files (*)", options=options)

        if file_name:
            try:
                # Чтение данных из Excel-файла
                df = pd.read_excel(file_name)

                # Заполнение полей ввода данными из Excel
                self.ui.entry_t0.setText(str(df["t0"].iloc[0]))
                self.ui.entry_tk.setText(str(df["tk"].iloc[0]))
                self.ui.entry_dt1.setText(str(df["dt1"].iloc[0]))
                self.ui.entry_dt2.setText(str(df["dt2"].iloc[0]))
                self.ui.entry_x0.setText(str(df["x0"].iloc[0]))
                self.ui.entry_v0.setText(str(df["v0"].iloc[0]))
                self.ui.entry_m.setText(str(df["m"].iloc[0]))
                self.ui.entry_k.setText(str(df["k"].iloc[0]))
                self.ui.entry_y.setText(str(df["y"].iloc[0]))

                # Проверка полей ввода и активация кнопки сброса
                self.check_input_fields()

            except Exception as e:
                # Отображаем сообщение об ошибке, если файл некорректен
                self.error_incorrect_input_message("Ошибка при чтении файла", f"Проверьте корректность "
                                                                              f"формата ввода данных в файле.")


if __name__ == "__main__":
    app = QApplication(sys.argv)  # Создаем объект приложения

    main_window = MainWindow()  # Создаем главное окно приложения
    main_window.show()  # Показываем главное окно

    sys.exit(app.exec())  # Запускаем основной цикл приложения и завершаем его при закрытии
